import calendar
from datetime import datetime, date
import os
from typing import Optional
import pandas as pd

from fastapi import APIRouter, Depends, Query
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app.core.database import get_db
from app.core.security import require_admin_or_superadmin
from app.models.user import User
from app.models.attendance import AttendanceLog
from app.models.company import Location
from app.models.holiday import Holiday
from app.models.working_days import WorkingDaysConfig
from app.routers.attendance import is_user_expected_working_day
from app.routers.payroll import get_employee_monthly_salary
from app.services.gps_service import calculate_distance_meters

router = APIRouter(
    prefix="/export",
    tags=["Export"]
)


def style_worksheet(ws):
    # Enable gridlines
    ws.views.sheetView[0].showGridLines = True

    # Font definitions
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=10)

    # Fills
    header_fill = PatternFill(start_color="0F5132", end_color="0F5132", fill_type="solid") # Dark green
    zebra_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid") # Off-white
    
    # Borders
    thin_border = Border(
        left=Side(style='thin', color='E0E0E0'),
        right=Side(style='thin', color='E0E0E0'),
        top=Side(style='thin', color='E0E0E0'),
        bottom=Side(style='thin', color='E0E0E0')
    )

    ws.row_dimensions[1].height = 28

    # Style headers
    for col_num in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        
    # Style rows
    for row_num in range(2, ws.max_row + 1):
        ws.row_dimensions[row_num].height = 20
        is_even = row_num % 2 == 0
        row_fill = zebra_fill if is_even else None
        
        for col_num in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.font = data_font
            cell.border = thin_border
            if row_fill:
                cell.fill = row_fill
                
            # Alignment and format rules
            header_val = ws.cell(row=1, column=col_num).value or ""
            
            # Alignments based on headers
            if any(term in header_val for term in ["Date", "ID", "Time", "Status", "Policy"]):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif any(term in header_val for term in ["Days", "Hours", "Count"]):
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.number_format = "0.0"
            elif "Salary" in header_val or "Payout" in header_val:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = "₹#,##0.00"
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # Autofit column dimensions
    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            val_str = str(cell.value or '')
            header_val = ws.cell(row=1, column=cell.column).value or ""
            if "Salary" in header_val or "Payout" in header_val:
                val_str = "₹" + val_str + ".00"
            max_length = max(max_length, len(val_str))
        ws.column_dimensions[column_letter].width = max(max_length, 12) + 3


@router.get("/attendance/xlsx")
def export_attendance_excel(
    query_date: Optional[str] = Query(None),
    year: Optional[int] = Query(None),
    month: Optional[int] = Query(None),
    employee_id: Optional[str] = Query(None),
    current_user: User = Depends(require_admin_or_superadmin),
    db: Session = Depends(get_db)
):
    # Fetch base queries
    logs_query = db.query(
        AttendanceLog,
        User
    ).join(
        User,
        AttendanceLog.user_id == User.id
    )

    if query_date:
        try:
            parsed_date = datetime.strptime(query_date, "%Y-%m-%d").date()
            logs_query = logs_query.filter(AttendanceLog.date == parsed_date)
        except ValueError:
            pass
    elif year and month:
        import calendar
        num_days = calendar.monthrange(year, month)[1]
        logs_query = logs_query.filter(
            AttendanceLog.date >= date(year, month, 1),
            AttendanceLog.date <= date(year, month, num_days)
        )
    elif year:
        logs_query = logs_query.filter(
            AttendanceLog.date >= date(year, 1, 1),
            AttendanceLog.date <= date(year, 12, 31)
        )

    if employee_id:
        from uuid import UUID
        is_uuid = False
        try:
            UUID(employee_id)
            is_uuid = True
        except ValueError:
            pass

        if is_uuid:
            logs_query = logs_query.filter(
                (User.employee_id == employee_id) | (User.id == employee_id)
            )
        else:
            logs_query = logs_query.filter(User.employee_id == employee_id)

    records = logs_query.order_by(
        AttendanceLog.date.desc(),
        AttendanceLog.checkin_time.desc()
    ).all()

    locations = db.query(Location).all()
    
    # Detailed log rows
    log_rows = []
    for attendance, user in records:
        matched_location = None
        if attendance.checkin_lat is not None and attendance.checkin_lng is not None:
            for loc in locations:
                dist = calculate_distance_meters(
                    attendance.checkin_lat,
                    attendance.checkin_lng,
                    loc.latitude,
                    loc.longitude
                )
                if dist <= loc.radius_meters + 10:
                    matched_location = loc.name
                    break
            if not matched_location and locations:
                closest_loc = min(
                    locations,
                    key=lambda l: calculate_distance_meters(
                        attendance.checkin_lat,
                        attendance.checkin_lng,
                        l.latitude,
                        l.longitude
                    )
                )
                matched_location = closest_loc.name

        def format_time(t):
            return t.strftime("%I:%M %p") if t else "-"

        notes_list = []
        if attendance.checkin_note:
            notes_list.append(f"In: {attendance.checkin_note}")
        if attendance.checkout_note:
            notes_list.append(f"Out: {attendance.checkout_note}")
        notes_str = " | ".join(notes_list) if notes_list else "-"

        moods = []
        if attendance.checkin_mood:
            mood_note_str = f" ({attendance.checkin_mood_note})" if attendance.checkin_mood_note else ""
            moods.append(f"In: {attendance.checkin_mood}{mood_note_str}")
        if attendance.checkout_mood:
            mood_note_str = f" ({attendance.checkout_mood_note})" if attendance.checkout_mood_note else ""
            moods.append(f"Out: {attendance.checkout_mood}{mood_note_str}")
        moods_str = " | ".join(moods) if moods else "-"

        log_rows.append({
            "Date": attendance.date.strftime("%Y-%m-%d") if attendance.date else "-",
            "Employee ID": user.employee_id,
            "Employee Name": user.name,
            "Email": user.email,
            "Check-in Time": format_time(attendance.checkin_time),
            "Check-out Time": format_time(attendance.checkout_time),
            "Check-in Status": attendance.checkin_status.replace("_", " ").title() if attendance.checkin_status else "-",
            "Check-out Status": attendance.checkout_status.replace("_", " ").title() if attendance.checkout_status else "-",
            "Total Hours": attendance.total_hours if attendance.total_hours is not None else 0.0,
            "Day Status": attendance.day_status.replace("_", " ").title() if attendance.day_status else "-",
            "Location Zone": matched_location or "Remote",
            "Notes": notes_str,
            "Mood Details": moods_str
        })

    os.makedirs("exports", exist_ok=True)
    file_name = f"attendance_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    file_path = os.path.join("exports", file_name)

    # Monthly summary logic (triggers when querying specifically for a month & year)
    if year and month:
        from app.models.holiday import Holiday
        from app.models.working_days import WorkingDaysConfig
        from app.models.comp_off import CompOffBalance, CompOffTransaction
        from app.utils.timezone import now_ist

        num_days = calendar.monthrange(year, month)[1]
        start_dt = date(year, month, 1)
        end_dt = date(year, month, num_days)

        working_days_cfg = db.query(WorkingDaysConfig).first()
        days_map = [True, True, True, True, True, True, False]
        if working_days_cfg:
            days_map = [
                working_days_cfg.monday, working_days_cfg.tuesday, working_days_cfg.wednesday,
                working_days_cfg.thursday, working_days_cfg.friday, working_days_cfg.saturday,
                working_days_cfg.sunday
            ]

        month_holidays = db.query(Holiday).filter(
            Holiday.date >= start_dt,
            Holiday.date <= end_dt
        ).all()
        holiday_dates = {h.date for h in month_holidays}
        current_date_ist = now_ist().date()

        # Fetch employees
        emp_query = db.query(User).filter(User.email != "admin@glrattendance.com")
        if employee_id:
            from uuid import UUID
            is_uuid = False
            try:
                UUID(employee_id)
                is_uuid = True
            except ValueError:
                pass

            if is_uuid:
                emp_query = emp_query.filter((User.employee_id == employee_id) | (User.id == employee_id))
            else:
                emp_query = emp_query.filter(User.employee_id == employee_id)
        
        employees = emp_query.order_by(User.name).all()
        summary_rows = []

        for emp in employees:
            logs = db.query(AttendanceLog).filter(
                AttendanceLog.user_id == emp.id,
                AttendanceLog.date >= start_dt,
                AttendanceLog.date <= end_dt
            ).all()
            log_by_date = {log.date: log for log in logs}

            expected_working_days = 0.0
            worked_days = 0.0
            paid_leaves = 0.0
            extra_days_worked = 0.0
            total_deductions = 0.0
            user_policy = emp.saturday_policy or "alt_sat_holiday"

            for d_num in range(1, num_days + 1):
                d = date(year, month, d_num)
                is_expected_work = is_user_expected_working_day(
                    d, user_policy, holiday_dates, days_map
                )
                log = log_by_date.get(d)

                if is_expected_work:
                    expected_working_days += 1.0
                    if log:
                        if log.day_status in ["full_day", "holiday_work"]:
                            worked_days += 1.0
                        elif log.day_status == "half_day":
                            worked_days += 0.5
                            total_deductions += 0.5
                        elif log.day_status == "comp_off_leave":
                            paid_leaves += 1.0
                        elif log.day_status == "absent":
                            total_deductions += 1.0
                    else:
                        if d <= current_date_ist:
                            total_deductions += 1.0
                else:
                    if log:
                        if log.day_status in ["full_day", "holiday_work"]:
                            worked_days += 1.0
                            extra_days_worked += 1.0
                        elif log.day_status == "half_day":
                            worked_days += 0.5
                            extra_days_worked += 0.5
                        elif log.day_status == "comp_off_leave":
                            paid_leaves += 1.0

            total_paid_days = max(0.0, 30.0 - total_deductions + extra_days_worked)
            base_salary = get_employee_monthly_salary(db, emp.id, year, month, emp.base_salary)
            calculated_salary = 0.0
            if base_salary > 0:
                calculated_salary = ((base_salary / 30.0) * total_paid_days) * 0.99

            total_hours_worked = sum(log.total_hours or 0.0 for log in logs)
            present_target_hours = worked_days * 9.0
            total_target_hours = expected_working_days * 9.0
            total_days_present = sum(1 for log in logs if log.checkin_time is not None and log.day_status != "absent")

            summary_rows.append({
                "Employee ID": emp.employee_id,
                "Employee Name": emp.name,
                "Email": emp.email,
                "Saturday Policy": user_policy.replace("_", " ").title(),
                "Expected Working Days Count": expected_working_days,
                "Total Target Hours (Expected Working Days x 9)": total_target_hours,
                "Total No of Days Present": total_days_present,
                "Present Days Count": worked_days,
                "Present Target Hours (Present Days x 9)": present_target_hours,
                "Actual Hours Worked": round(total_hours_worked, 2),
                "Paid Leaves": paid_leaves,
                "Extra Days Worked (Holiday/Weekend)": extra_days_worked,
                "Net Paid Days (30-day billing)": total_paid_days,
                "Base Monthly Salary": base_salary,
                "Calculated Net Salary (99%)": round(calculated_salary, 2)
            })

        df_summary = pd.DataFrame(summary_rows)
        df_logs = pd.DataFrame(log_rows)

        if df_summary.empty:
            df_summary = pd.DataFrame(columns=[
                "Employee ID", "Employee Name", "Email", "Saturday Policy",
                "Expected Working Days Count", "Total Target Hours (Expected Working Days x 9)",
                "Total No of Days Present", "Present Days Count", "Present Target Hours (Present Days x 9)",
                "Actual Hours Worked", "Paid Leaves", "Extra Days Worked (Holiday/Weekend)",
                "Net Paid Days (30-day billing)", "Base Monthly Salary", "Calculated Net Salary (99%)"
            ])
        if df_logs.empty:
            df_logs = pd.DataFrame(columns=[
                "Date", "Employee ID", "Employee Name", "Email", "Check-in Time", "Check-out Time",
                "Check-in Status", "Check-out Status", "Total Hours", "Day Status", "Location Zone",
                "Notes", "Mood Details"
            ])

        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            df_summary.to_excel(writer, index=False, sheet_name="Monthly Summary")
            df_logs.to_excel(writer, index=False, sheet_name="Daily Logs Details")

            style_worksheet(writer.sheets["Monthly Summary"])
            style_worksheet(writer.sheets["Daily Logs Details"])

    else:
        # Simple report (not monthly filter)
        df_logs = pd.DataFrame(log_rows)
        if df_logs.empty:
            df_logs = pd.DataFrame(columns=[
                "Date", "Employee ID", "Employee Name", "Email", "Check-in Time", "Check-out Time",
                "Check-in Status", "Check-out Status", "Total Hours", "Day Status", "Location Zone",
                "Notes", "Mood Details"
            ])

        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            df_logs.to_excel(writer, index=False, sheet_name="Daily Logs Details")
            style_worksheet(writer.sheets["Daily Logs Details"])

    return FileResponse(
        path=file_path,
        filename=file_name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )